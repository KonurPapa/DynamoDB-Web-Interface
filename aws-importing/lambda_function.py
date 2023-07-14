import dropbox
import boto3
import io
import json
import openpyxl
from openpyxl import Workbook
import openpyxl.utils as utils
from openpyxl.utils import column_index_from_string
import re as regex
import time
from datetime import datetime
from decimal import Decimal

dynamodb = boto3.resource("dynamodb")


def lambda_handler(event, context):
    
    # DynamoDB database to import into
    table = "Landscape_Test"
    
    # Name of the sheet within the Excel file to import data from
    sheet_name = "Project Totals"
    
    # App key and secret
    app_key = "123abc456def"
    app_secret = "abc123xyz789"
    
    # Refresh token, which will perpetually generate new sessions
    refresh_token = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
    
    # Create a Dropbox client using the authorizer
    dbx = dropbox.Dropbox(app_key=app_key, app_secret=app_secret, oauth2_refresh_token=refresh_token)
    
    # List the contents of the folder (no path because the root is the directory we want)
    result = dbx.files_list_folder("")
    
    # Whether to strip out subheader numbers on import
    strip_header_nums = True
    
    
    # --- WRITING DATA ---
    
    # List of all sheet files
    file_list = []
    
    # Console log sheet data
    out_data = {
        "all_files": [],
        "all_rows": [],
        "all_time": [],
        
        "imported": [],
        "imported_rows": [],
        "imported_time": [],
        
        "lost": [],
        "lost_errors": [],
        "lost_time": [],
        
        "time_start": 0,
        "time_end": 0
    }
    
    # Loop through the results and retrieve metadata for each file
    while True:
        for entry in result.entries:
            # Only add files to the list
            if isinstance(entry, dropbox.files.FileMetadata):
                file_list.append(entry)
        # If there are more results, continue to the next page
        if result.has_more:
            result = dbx.files_list_folder_continue(result.cursor)
        else:
            break
    
    # Cells class: stores cell selection data
    class Cells:
        def __init__(self, start_cell, end_cell):
            self.start_cell = start_cell
            self.end_cell = end_cell
            
            self.start_col = column_index_from_string(regex.findall('[A-Z]+', start_cell)[0])
            self.start_row = int(regex.findall('[0-9]+', start_cell)[0])
            self.end_col = column_index_from_string(regex.findall('[A-Z]+', end_cell)[0])
            self.end_row = int(regex.findall('[0-9]+', end_cell)[0])
            
            self.desc_col = 1
            
            self.selection = sheet[self.start_cell:self.end_cell]
        
        def update_sel(self):
            self.start_cell = utils.get_column_letter(self.start_col) + str(self.start_row)
            self.end_cell = utils.get_column_letter(self.end_col) + str(self.end_row)
            
            self.selection = sheet[self.start_cell:self.end_cell]
    
    # Loop through all returned files
    for index in file_list:
        file = index.name
        out_data["time_start"] = time.perf_counter()
        
        # Download file from Dropbox to temp directory
        if file.endswith(".xls") or file.endswith(".xlsx") or file.endswith(".xlsm"):
            download = dbx.files_download_to_file("/tmp/temp_file.xlsx", "/" + file)
        
        # Create pretty sheet name (get project name from file name)
        # match = regex.search(r"(?<= - ).+(?= - )", file)
        match = regex.search(r"^.+(?= - )", file)
        if match:
            name = match.group()
        else:
            name = "Untitled Project"
        
    
        # --- GETTING / READING DATA ---
        
        # Open the workbook
        wb = openpyxl.load_workbook("/tmp/temp_file.xlsx")
        
        # Select the proper worksheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
        
        # Create Cell objects with cell coordinates (manually specified)
        meta = Cells("A1", "B1")
        db = Cells("A2", "B2")
        
        
        # --- FINDING METADATA BOUNDS ---
        
        try:
            # Finding meta start row, in case there are empty rows/header text
            for row in range(1, sheet.max_row):
                if sheet.cell(row=row, column=1).value is not None and "sheet" not in sheet.cell(row=row, column=1).value.lower():
                    meta.start_row = row
                    break
            
            # Set meta end row, based on how many metadata headers there are
            for i, cell in enumerate(sheet.iter_rows(min_row=meta.start_row, max_row=sheet.max_row, min_col=1, max_col=1)):
                if cell[0].value is None:
                    meta.end_row = i
                    break
                # Clean up header formatting
                cell[0].value = cell[0].value.replace(":", "").rstrip().lstrip()
            
            
            # --- FINDING DB DATA BOUNDS ---
            
            # Find start row: Starting at the end of the metadata, check for the first row with a value that's bold (a header) and doesn't use _ (which would signify it's a macro line)
            for row in range(meta.end_row + 1, sheet.max_row):
                if sheet.cell(column=2, row=row).value is not None and sheet.cell(column=2, row=row).font.bold and "_" not in sheet.cell(column=2, row=row).value:
                    db.start_row = row
                    break
            
            # Find description column from the header row
            for col in range(1, sheet.max_column):
                desc = ["Description", "Item"]
                if sheet.cell(row=db.start_row, column=col).value is not None and sheet.cell(row=db.start_row, column=col).value.lower() == (desc[0].lower() or desc[1].lower()):
                    # Rename to "Description"
                    if sheet.cell(row=db.start_row, column=col).value.lower() is not desc[0]:
                        sheet.cell(row=db.start_row, column=col).value = desc[0]
                    db.desc_col = col
                    break
            
            # Find end row: Check if current cell and following is empty, and if it is, check the same cell and following in the Type column to the left of it (after making sure it exists)
            # for row in range(db.start_row, sheet.max_row + 1):
            #     if ((sheet.cell(row=row, column=db.desc_col).value is None) and sheet.cell(row=row + 1, column=db.desc_col) is None) or row >= sheet.max_row:
            #         if db.desc_col > 1:
            #             if sheet.cell(row=row, column=db.desc_col - 1).value is None and sheet.cell(row=row + 1, column=db.desc_col - 1).value is None:
            #                 db.end_row = row
            #                 break
            #         else:
            #             db.end_row = row
            #             break
            
            # Check for 2 consecutive empty rows, and if found, set the row end
            for row in range(db.start_row, sheet.max_row + 1):
                if (sheet.cell(row=row, column=db.desc_col).value is None and sheet.cell(row=row + 1, column=db.desc_col).value is None) or row >= sheet.max_row:
                    # If there's another column to the left of Description, check to make sure it's empty too (sometimes headers will be in this column)
                    if db.desc_col > 1:
                        if sheet.cell(row=row, column=db.desc_col - 1).value is None and sheet.cell(row=row + 1, column=db.desc_col - 1).value is None:
                            db.end_row = row
                    else:
                        db.end_row = row
            
            # Find start column: Shift start column forward if column A is empty (most sheets start db data on B)
            if all(cell[0].value is None for cell in sheet.iter_rows(min_row=db.start_row, max_row=db.end_row, min_col=db.start_col, max_col=db.start_col)):
                db.start_col += 1
            
            # Find end column
            for col in range(db.start_col, sheet.max_column + 1):
                if sheet.cell(row=db.start_row, column=col).value is None or sheet.column_dimensions[sheet.cell(row=db.start_row, column=col).column_letter].hidden:
                    db.end_col = col - 1
                    break
            
            # Update cell selection data with the coord modifications above
            meta.update_sel()
            db.update_sel()
            
            
            # --- REFORMATTING DB DATA ---
            
            # Create lists of subheaders to be appended, and keep track of the row offset each time we find one
            head_type = []
            head_subtype = []
            type_offset = 0
            subtype_offset = 0
            
            # Finding subheaders in first data column (start 1 row down to ignore header)
            for i, cell in enumerate(sheet.iter_rows(min_row=db.start_row + 1, max_row=db.end_row, min_col=db.start_col, max_col=db.start_col)):
                if cell[0].font.bold and cell[0].value is not None:
                    # Organize by type/subtype, saving values and cell indexes for data population. Offset by how many empty subheader cells we've created (since we can't actually delete them yet)
                    if (" - " in cell[0].value and cell[0].font.bold) or cell[0].value.isupper():
                        subtype_offset += 1
                        
                        # Strip leading number out of subheader
                        if " - " in cell[0].value and strip_header_nums:
                            cell[0].value = cell[0].value[4:]
                        
                        cell[0].value = cell[0].value.rstrip().lstrip()
                        head_subtype.append([cell[0].value.title() if not " - " in cell[0].value else cell[0].value, i - subtype_offset])
                    else:
                        type_offset += 1
                        
                        cell[0].value = cell[0].value.rstrip().lstrip()
                        head_type.append([cell[0].value, i - type_offset])
                    
                    cell[0].value = None
            
            # Delete the first data column if it's now empty (some sheets have only subheaders saved in first column, which at this point have been moved)
            if all(cell[0].value is None for cell in sheet.iter_rows(min_row=db.start_row + 1, max_row=db.end_row, min_col=db.start_col, max_col=db.start_col)):
                db.start_col += 1
                db.update_sel()
            
            
            # --- TRANSPOSING & APPENDING DATA ---
            
            # Transpose the metadata to iterate by column instead of row
            meta_flip = zip(*meta.selection)
            meta_cols = []
            
            # Iterate over transposed cells by column
            for col_data in meta_flip:
                meta_cols.append([cell.value.strftime("%Y-%m-%d") if isinstance(cell.value, datetime) else cell.value.rstrip().lstrip() if isinstance(cell.value, str) else cell.value for cell in list(col_data)])
            
            # Retrieve the values from each cell and convert to 2-dimensional array of data
            data = [[cell.value.strftime("%Y-%m-%d") if isinstance(cell.value, datetime) else cell.value for cell in row] for row in db.selection]
            
            # Delete any empty rows from db data
            for row in range(len(data) - 1, 0, -1):
                if data[row][0] is None:
                    data.pop(row)
            
            # Appending type and subtype headers and data to each row, and metadata info
            type_iter = 0
            subtype_iter = 0
            unique_keys = {}
            for i, row in enumerate(data):
                if i == 0:
                    row += ["Type", "Sub-Type", "Serial"]
                    row += meta_cols[0]
                else:
                    # Only iterate type/subtype index if another header exists in the list and we've hit its iterate limit
                    if type_iter + 1 < len(head_type):
                        if i >= head_type[type_iter + 1][1]:
                            type_iter += 1
                    
                    if subtype_iter + 1 < len(head_subtype):
                        if i >= head_subtype[subtype_iter + 1][1]:
                            subtype_iter += 1
                    
                    # Assign unique item number to serial
                    if row[0] in unique_keys:
                        unique_keys[row[0]] += 1
                    else:
                        unique_keys[row[0]] = 1
                    
                    # Append headers, serial, metadata
                    row += [head_type[type_iter][0], head_subtype[subtype_iter][0], "".join(regex.findall(r"[A-Z]", name)).replace(" ", "")[:3] + "-" + "".join(regex.findall(r"[A-Z]|[0-9]", row[0])).replace(" ", "")[:4] + "-" + str(unique_keys[row[0]])]
                    row += meta_cols[1]
        
        # If there's a problem finding data bounds, log the error and continue
        except Exception as err:
            out_data["lost_errors"].append(err)
            out_data["lost"].append(name)
            continue
        
        
        # --- IMPORTING DATA TO DYNAMODB ---
        
        # Log an error if no sheet data is returned (or if the default of 2 is found, meaning the sheet data length hasn't changed)
        if len(data) <= 2:
            out_data["lost_errors"].append("Sheet data not found")
            out_data["lost"].append(name)
            continue
        
        batch_items = []
        
        # Reformat data into JSON key/value pairs for importing into DynamoDB
        try:
            for row in data[1:]:
                item = {}
                for i, value in enumerate(row):
                    if value is not None:
                        item[data[0][i]] = value
                        
                        # Converting decimal values (needs testing)
                        # item[data[0][i]] = regex.sub(r"\b\d+\.\d+\b", str(Decimal(value)), item[data[0][i]])
                
                # dynamodb.Table(table).put_item(Item=item)
                batch_items.append({'PutRequest': {'Item': item}})
        except TypeError as err:
            continue
        
        # Split the batch items into chunks of 25 (DynamoDB batch limit)
        batch_chunks = [batch_items[i:i + 25] for i in range(0, len(batch_items), 25)]
        
        # Execute the batch_write_item() request for each chunk
        for chunk in batch_chunks:
            response = dynamodb.batch_write_item(
                RequestItems = {
                    table: chunk
                }
            )
        
            # Check for unprocessed items
            unprocessed_items = response.get('UnprocessedItems', {})
            
            # Log processed/unprocessed items
            if unprocessed_items:
                out_data["lost_rows"].append(len(unprocessed_items))
                out_data["imported_rows"].append((db.end_row - db.start_row) - len(unprocessed_items))
            else:
                out_data["imported_rows"].append(db.end_row - db.start_row)
        
        # Add name to human-readable list
        out_data["imported"].append(name)
        out_data["time_start"] = time.perf_counter()
        out_data["all_time"] = int(out_data["time_end"]) - int(out_data["time_start"])
    
    return {
        "INFO": str(len(out_data["imported"])) + " sheets imported successfully. " + str(len(out_data["lost"])) + " sheets failed to import.",
        "ALL FILES": str(out_data["all_files"]) + ". Number of rows found per sheet: " + str(out_data["all_rows"]) + ". Runtime per sheet: " + str(out_data["all_time"]),
        "SUCCESS": str(out_data["imported"]) + ". Number of rows imported per sheet: " + str(out_data["imported_rows"]) + ". Runtime per sheet: " + str(out_data["imported_time"]),
        "FAIL": str(out_data["lost"]) + ". Errors logged for sheets: " + str(out_data["lost_errors"]) + ". Runtime per sheet: " + str(out_data["lost_time"])
    }
